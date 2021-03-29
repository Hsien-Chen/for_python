# nl=[number*0.01 for number in range(1,6)]
# print(nl)
# count= 0
# rate = int(input('rate:'))
# while count < rate:
#     print(count)
#     count+=1
from openpyxl import Workbook
import numpy as np
from openpyxl.chart import (
    AreaChart,
    Reference,
    Series,
)
lv=float(input('貸款金額:'))
repay='本息攤還' #input('還款方式:')
yr=int(input('貸款年數:'))
yrate=float(input('年利息:'))
 # 創建一個空白活頁簿物件
wb = Workbook()
# 選取正在工作中的表單
ws = wb.active
ws1=wb.create_sheet("本金")
# 指定值給 A1 儲存格
ws['A1'] = '貸款金額'
ws['A2'] = '還款方式'
ws['A3'] = '貸款年數'
ws['A4'] = '年利息'
ws['A5'] = '總繳利息'

ws['B1'] = lv
ws['B2'] = repay
ws['B3'] = yr
ws['B4'] = yrate


ws['A7'] = '期數(年)'
ws['B7'] = '貸款餘額'
ws['C7'] = '本金攤還'
ws['D7'] = '利息'
ws['E7'] = '繳款金額'
rf= (1+(yrate*0.01/12))**(yr*12)
rf1=rf*(yrate*0.01/12)
rf2=rf-1
rf3=(rf1)/(rf2)

ct=1
while ct <= yr*12:
    ws.cell(row=ct+8,column=1,value=str(ct))
    ws.cell(row=ct+8,column=5,value=rf3*lv)
    ct += 1 

ws['B8']=lv
ws['D9']=lv*2.5*0.01/12
ws['C9']=ws['E9'].value - ws['D9'].value
ws['B9']=ws['B8'].value - ws['C9'].value

for ct in range(1,yr*12):

    a=ws.cell(row=ct+8,column=2) #上一期貸款餘額
    b=a.value #上一期貸款餘額 int
    ws.cell(row=ct+9,column=4).value=b*yrate*0.01/12
     #利息
    
    d=ws.cell(row=ct+9,column=5).value #繳款金額
    ws.cell(row=ct+9,column=3).value=d - (b*yrate*0.01/12)#本期繳款本金
    aplusaone=ws.cell(row=ct+9,column=2)#本期貸款餘額
    aplusaone.value=b-d+(b*yrate*0.01/12)

allrate=(rf3*lv)*yr*12
ws['B5'] = allrate 
rows=ws['A9':'E248']
for row in ws.iter_rows(min_row=8,min_col=1,max_row=8+yr*12,max_col=5):
    for cell in row:
        cell

chart = AreaChart()
chart.title = "本息平均攤還"
chart.style = 13
chart.x_axis.title = '期數'
chart.y_axis.title = '金額'

cats = Reference(ws, min_col=1, min_row=9, max_row=9+yr*12)
data = Reference(ws, min_col=3, min_row=9, max_col=4, max_row=9+yr*12)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

ws.add_chart(chart, "I9")
   

#本金平均攤還
ws1['A1'] = '貸款金額'
ws1['A2'] = '還款方式'
ws1['A3'] = '貸款年數'
ws1['A4'] = '年利息'
ws1['A5'] = '總繳利息'

ws1['B1'] = lv
ws1['B2'] = repay
ws1['B3'] = yr
ws1['B4'] = yrate


ws1['A7'] = '期數(年)'
ws1['B7'] = '貸款餘額'
ws1['C7'] = '本金攤還'
ws1['D7'] = '利息'
ws1['E7'] = '繳款金額'
# rf= (1+(yrate*0.01/12))**(yr*12)
# rf1=rf*(yrate*0.01/12)
# rf2=rf-1
# rf3=(rf1)/(rf2)

ct=1
while ct <= yr*12:
    ws1.cell(row=ct+8,column=1,value=str(ct))
    ws1.cell(row=ct+8,column=3,value=lv/(yr*12))

    ct += 1 

ws1['B8']=lv
ws1['B9']=ws1['B8'].value - ws1['C9'].value
ws1['D9']=ws1['B9'].value *yrate*0.01/12
ws1['E9']=ws1['C9'].value +ws1['D9'].value
for ct in range(1,yr*12):

    a=ws1.cell(row=ct+8,column=2) #上一期貸款餘額
    b=a.value #上一期貸款餘額 int
    z=ws1.cell(row=ct+9,column=3).value #本期本金
    ws1.cell(row=ct+9,column=2).value=b-z #本期貸款餘額
    ws1.cell(row=ct+9,column=4).value=(b-z)*yrate*0.01/12
     #利息
    



    d=ws1.cell(row=ct+9,column=5,value=z+(b-z)*yrate*0.01/12) #繳款金額
    # ws1.cell(row=ct+9,column=3).value=d - (b*yrate*0.01/12)#本期繳款本金
    # aplusaone=ws1.cell(row=ct+9,column=2)#本期貸款餘額
    # aplusaone.value=b-d+(b*yrate*0.01/12)


# ws['B5'] = allrate 
# rows=ws1['A9':'E248']
# for row in ws1.iter_rows(min_row=8,min_col=1,max_row=8+yr*12,max_col=5):
#     for cell in row:
#         print(cell)

c1 = AreaChart()
c1.title = "本金平均攤還"
c1.style = 13
c1.x_axis.title = '期數'
c1.y_axis.title = '金額'

cs = Reference(ws1, min_col=1, min_row=9, max_row=9+yr*12)
data = Reference(ws1, min_col=3, min_row=9, max_row=9+yr*12)
data1 = Reference(ws1, min_col=5, min_row=9, max_row=9+yr*12)
chart.add_data(data, titles_from_data=True,from_rows=True)
chart1.add_data(data1, titles_from_data=True,from_rows=True)
chart.set_categories(cats)

ws1.add_chart(chart, "I9")




# 儲存成 create_sample.xlsx 檔案
wb.save('creat.xlsx')